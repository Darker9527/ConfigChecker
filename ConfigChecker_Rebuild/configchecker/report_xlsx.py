# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from .models import CheckTask, CheckResult


def write_xlsx_report(task: CheckTask, results: list[CheckResult], out_dir: str | Path) -> Path:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f'{task.name}_{task.id or "new"}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = '检查结果'
    header = ['检查项', '风险', '状态', '证据', '原因', '整改建议']
    ws.append(['任务名称', task.name, '设备类型', task.devtype, '设备品牌', task.brand])
    ws.append(['配置文件', task.config_path])
    ws.append([])
    ws.append(header)

    fill = PatternFill('solid', fgColor='1F4E79')
    font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[4]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for r in results:
        ws.append([r.baseline_name, r.risk, r.status, r.evidence, r.reason, r.suggestion])
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    widths = [35, 12, 12, 50, 40, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w
    ws.freeze_panes = 'A5'

    ws2 = wb.create_sheet('汇总')
    summary = {}
    risk_summary = {}
    for r in results:
        summary[r.status] = summary.get(r.status, 0) + 1
        risk_summary[r.risk] = risk_summary.get(r.risk, 0) + 1
    ws2.append(['状态', '数量'])
    for k, v in summary.items():
        ws2.append([k, v])
    ws2.append([])
    risk_start = ws2.max_row + 1
    ws2.append(['风险', '数量'])
    for risk in ['高风险', '中风险', '低风险']:
        ws2.append([risk, risk_summary.get(risk, 0)])
    chart = BarChart()
    chart.title = '风险统计'
    chart.y_axis.title = '数量'
    chart.x_axis.title = '风险等级'
    data = Reference(ws2, min_col=2, min_row=risk_start, max_row=risk_start + 3)
    cats = Reference(ws2, min_col=1, min_row=risk_start + 1, max_row=risk_start + 3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws2.add_chart(chart, 'D2')
    wb.save(path)
    return path
