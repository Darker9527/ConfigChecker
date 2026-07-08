# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import CheckTask, CheckResult

RISK_ORDER = {'高风险': 0, '中风险': 1, '低风险': 2}
RISK_FILL = {'高风险': 'FEE2E2', '中风险': 'FEF3C7', '低风险': 'DCFCE7'}


def write_remediation_xlsx(task: CheckTask, results: list[CheckResult], out_dir: str | Path, only_failed: bool = True) -> Path:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f'{task.name}_{task.id or "new"}_整改台账.xlsx'
    rows = [r for r in results if (not only_failed or r.status == '未通过')]
    rows.sort(key=lambda r: (RISK_ORDER.get(r.risk, 9), r.baseline_name))

    wb = Workbook()
    ws = wb.active
    ws.title = '整改台账'
    ws.append(['任务名称', task.name, '设备类型', task.devtype, '设备品牌', task.brand])
    ws.append(['配置文件', task.config_path, '待整改项', len(rows)])
    ws.append([])
    headers = ['序号', '风险', '检查项', '状态', '问题原因', '整改建议', '证据', '责任人', '计划完成时间', '整改状态', '复测结果', '备注']
    ws.append(headers)

    head_fill = PatternFill('solid', fgColor='1F4E79')
    head_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[4]:
        cell.fill = head_fill
        cell.font = head_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, r in enumerate(rows, 1):
        ws.append([i, r.risk, r.baseline_name, r.status, r.reason, r.suggestion, r.evidence, '', '', '未整改', '', ''])
        row = ws.max_row
        fill = PatternFill('solid', fgColor=RISK_FILL.get(r.risk, 'FFFFFF'))
        ws.cell(row=row, column=2).fill = fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical='top')

    widths = [8, 12, 36, 12, 42, 48, 48, 14, 16, 14, 14, 24]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=idx).column_letter].width = width
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:L{max(4, ws.max_row)}'

    summary = wb.create_sheet('风险汇总')
    summary.append(['风险', '数量'])
    counts = {}
    for r in rows:
        counts[r.risk] = counts.get(r.risk, 0) + 1
    for risk in ['高风险', '中风险', '低风险']:
        summary.append([risk, counts.get(risk, 0)])
    wb.save(path)
    return path
